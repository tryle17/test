import os
import argparse
import re
import sys
import subprocess
import shutil
from datetime import datetime
import requests
import json
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import urllib3
from requests.auth import HTTPBasicAuth

file_paths = [
    ("boot.img", ""),
    ("boot-lz4.img", ""),
    ("boot-gz.img", ""),
    ("system_dlkm.flatten.erofs.img", "system_dlkm.img"),
    ("vmlinux", ""),
    ("gki-info.txt", ""),
    ("vmlinux.symvers", "")
]

repo_cmd = "repo"
base_path = ""

def get_tools_path(relative_path):
    global base_path

    if not base_path:
        base_path = os.path.dirname(os.path.realpath(__file__))
    full_path = os.path.join(base_path, relative_path)
    print("base_path {} relative_path {} full_path {}".format(base_path, relative_path, full_path))
    return full_path

def extract_android_version(branch):
    pattern = r'android\d+-\d+(\.\d+)?'
    match = re.search(pattern, branch)
    if match:
        return match.group()
    else:
        return None

def get_current_date():
    now = datetime.now()
    return now.strftime('%Y-%m')

def generate_build_file_name(branch, date, release, number):
    return "{}-{}_r{}_abogki{}".format(branch, date, release, number)

def extract_and_increment(text):
    pattern = r'r(\d+)'
    match = re.search(pattern, text)

    if match:
        result = int(match.group(1)) + 1
        #result = int(match.group(1))
        return result
    else:
        return None

def process_config_file(config):
    if not os.path.exists(config):
        print(f"The file {config} does not exist.")
        return 1

    patterns = [
        re.compile(r"android\d+-\d+\.\d+-\d+-\d+_r\d+\b"),
        re.compile(r"android\d+-\d+\.\d+-\d+-\d+_r\d+_abogki\d+\b")
    ]
    results = []

    try:
        with open(config, "r") as file:
            for line in file:
                line = line.strip()  # 去除前后空白字符
                for pattern in patterns:
                    match = pattern.search(line)
                    #print('pattern:', pattern)
                    #print('line:', line)
                    #print('match:', match)
                    if match:
                        result = extract_and_increment(match.group())
                        results.append(result)
    except Exception as e:
        print(f"Error processing the file {config}: {e}")
        return 1

    if results:
        return max(results)
    else:
        return 1

def get_current_release(out):
    command = [
        "python3", get_tools_path("ogki_gki_artifactory.py"),
        "-t", "get",
        "-k", "OGKI",
        "-c", "release.txt",
        "-i", out
    ]
    print('\ncommand',command)
    print('\n')

    try:
        result = subprocess.run(command, shell=False, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    except subprocess.CalledProcessError as e:
        print("run command error： {}".format(e))
    release = process_config_file(os.path.join(out, "release.txt"))
    print('\ncurrent_release',release)
    #sys.exit(1)
    return release

def check_if_partner(input_str):
    if "partner-android" in input_str:
        return "partner"
    else:
        return "aosp"

def process_path(path, type, branch):
    # 去除路径最后的'/'
    path = path.rstrip('/')
    # 获取路径的最后一个目录名
    last_dir = os.path.basename(path)
    parent_dir = os.path.dirname(path)

    # 检查并追加type变量的值
    if type not in last_dir:
        last_dir += "-{}".format(type)

    # 检查并追加branch变量的值
    if branch not in last_dir:
        last_dir += "-{}".format(branch)

    # 生成新路径
    new_path = os.path.join(parent_dir, last_dir)
    return new_path

def create_symlink(src, dst):
    if not os.path.exists(src):
        print("Source directory does not exist:", src)
        return
    if os.path.exists(dst):
        print("Destination directory already exists:", dst)
        return
    os.symlink(src, dst)
    print("Soft link created successfully:", dst, "->", src)

def is_token_valid(token):
    """
    验证 token 是否为一个有效的数字串
    不能有/ 或者 其他字符
    """
    return token.isdigit()

def parse_gerrit_number(input_str):
    """
    解析输入字符串，提取最后两个数字并返回特定格式化字符串，以及状态。

    Args:
        input_str (str): 输入的字符串。

    Returns:
        tuple: 格式化字符串和状态（True/False）。
    """
    # 使用正则表达式找到所有数字
    all_tokens = re.split(r'/+', input_str)

    # 筛选出有效的数字token
    valid_numbers = [token for token in all_tokens if is_token_valid(token)]

    if not valid_numbers:
        return "", "", False

    last_two_numbers = valid_numbers[-2:]

    if len(last_two_numbers) == 1:
        id_number = last_two_numbers[0]
        return id_number, "", True

    id_number = last_two_numbers[0]
    current_revision_number = last_two_numbers[1]

    last_two_digits = id_number[-2:] if len(id_number) >= 2 else id_number

    formatted_string = "refs/changes/{}/{}/{}".format(last_two_digits, id_number, current_revision_number)

    return formatted_string, current_revision_number, False

def get_absolute_path(path):
    if os.path.isabs(path):
        # 如果是绝对路径，直接返回完整的绝对路径
        return os.path.abspath(path)
    else:
        # 如果是相对路径，将当前路径和相对路径拼接成绝对路径
        current_dir = os.getcwd()
        absolute_path = os.path.join(current_dir, path)
        return os.path.abspath(absolute_path)

def parse_cmd_args():

    parser = argparse.ArgumentParser(description="ogki gki oki build tools")
    parser.add_argument('-p', '--path', type=str, help='default dirname', default=get_absolute_path('google'))
    parser.add_argument('-u', '--url', type=str, help='manifest url',
           default='https://partner-android.googlesource.com/kernel/manifest')
    parser.add_argument('-s', '--common', type=str, help='aosp source code', default='')
    parser.add_argument('-b', '--branch', type=str, help='manifest branch', default='common-android15-6.6-oppo')
    parser.add_argument('-m', '--manifest', type=str, help='manifest file name', default='')
    parser.add_argument('-e', '--ext', type=str, help='repo special ext command', default='')
    parser.add_argument('-t', '--type', type=str, help='branch private or public', default='partner')
    parser.add_argument('-o', '--out', type=str, help='dist output', default='oplus')
    parser.add_argument('-z', '--zip', type=str, help='zip file name', default='dist.tar.gz')
    parser.add_argument('-n', '--bug_number', type=str, help='Google issue number', default='000000000')
    parser.add_argument('-g', '--gerrit_number', type=str, help='Google gerrit number', default='')
    parser.add_argument('-c', '--config_info', type=str, help='config info', default=get_tools_path('config_info.xlsx'))
    parser.add_argument('-r', '--review_url', type=str, help='review gerrit url', default='')
    parser.add_argument('-w', '--gerrit_url', type=str, help='web gerrit url', default='')
    parser.add_argument('-a', '--auto', action='store_true', help='auto get info', default=False)
    parser.add_argument('-l', '--log', type=str, help='out put log dir', default='')
    parser.add_argument('-x', '--max', type=str, help='max change list', default='10')
    parser.add_argument('--refs', type=str, help='refs change', default='')
    parser.add_argument('--out_abi', type=str, help='out abi', default='out_abi/kernel_aarch64/dist')
    parser.add_argument('--user', type=str, help='default user', default='')
    parser.add_argument('--new', type=str, help='new build', default='true')
    parser.add_argument('--artifactory', type=str, help='default user', default='')

    args = parser.parse_args()

    args.path = get_absolute_path(args.path)
    args.type = check_if_partner(args.url)
    branch = extract_android_version(args.branch)
    current_path = process_path(args.path, args.type, branch) # 输出：google-aosp-android15-6.6
    print("current_path:", current_path)
    create_symlink(args.path, current_path)
    args.path = current_path
    date = get_current_date()

    tmp_path = generate_build_file_name(branch,date, "0", args.bug_number)
    tmp_full_path = os.path.join(args.path, tmp_path)

    final_path = generate_build_file_name(branch,date, get_current_release(tmp_full_path), args.bug_number)
    final_full_path = os.path.join(args.path, final_path)

    if not os.path.exists(final_full_path) and os.path.exists(tmp_full_path):
        os.rename(tmp_full_path, final_full_path)
    elif os.path.exists(tmp_full_path):
        shutil.rmtree(tmp_full_path)
    args.out = os.path.join(args.path, final_path)
    args.zip = os.path.join(args.out, final_path + ".tar.gz")
    args.common = os.path.join(args.path, "common")
    args.log = os.path.join(args.out, "log")
    args.out_abi = os.path.join(args.path, "out_abi/kernel_aarch64/dist")

    if not args.artifactory:
        args.artifactory =  args.out

    if args.type.lower() == "partner":
        args.review_url = "https://partner-android-review.googlesource.com"
        args.gerrit_url = "https://partner-android.googlesource.com/kernel-oem/oppo/common"
    else:
        args.review_url = "https://android-review.googlesource.com"
        args.gerrit_url = "https://android.googlesource.com/kernel/common"

    args.refs, revision_number, args.auto = parse_gerrit_number(args.gerrit_number)
    #print("input {}\nparse：{}\nrebase state：{}\n".format(args.gerrit_number, args.refs, args.auto))

    print('path:', args.path, '\nurl:', args.url,
      '\ncommon:', args.common,'\nbranch:', args.branch,
      '\nmanifest:', args.manifest, '\next command:', args.ext,
      '\ntype:', args.type,  '\nout:', args.out,  '\nzip:', args.zip,
      '\nbug_number:', args.bug_number, '\ngerrit_number:', args.gerrit_number,
      '\nconfig_info:', args.config_info,
      '\nreview_url:', args.review_url, '\ngerrit_url:', args.gerrit_url,
      '\nauto:', args.auto, '\nlog:', args.log,
      '\nmax:', args.max,' \nrefs:', args.refs,' \nout_abi:', args.out_abi,
      '\nnew:', args.new,'\nartifactory:', args.artifactory)

    if not args.gerrit_number:
        print("now your gerrit number is null,you must input a gerrit number like 3249938 or 3249938/1")
        sys.exit(1)

    if not revision_number and args.type.lower() == "partner":
        message = (
            "Partner Gerrit is private, cannot find revision number by command. \n"
            "You must input the full form like '3249938/1'. The form '3249938' is not supported."
        )
        print(message)
        sys.exit(1)
    return args

def delete_matching_directories(base_directory, patterns):
    """
    在基目录中查找匹配指定正则表达式模式的所有目录，
    如果找到匹配的目录，则删除它们。

    :param base_directory: 要检查的基目录路径
    :param patterns: 模糊匹配的正则表达式模式列表
    """
    if not os.path.isdir(base_directory):
        print(f"The directory {base_directory} does not exist or is not a directory.")
        return

    # 列出基目录中的所有子目录
    subdirectories = next(os.walk(base_directory))[1]

    # 记录所有匹配的目录
    matched_directories = []

    for subdir_name in subdirectories:
        subdir_path = os.path.join(base_directory, subdir_name)

        for pattern in patterns:
            if re.search(pattern, subdir_name):
                matched_directories.append(subdir_path)
                # 不需要 break，这样会继续检查其他模式

    # 删除所有匹配的目录
    for dir_path in matched_directories:
        try:
            shutil.rmtree(dir_path)
            print(f"Deleted directory: {dir_path}")
        except Exception as e:
            print(f"Failed to delete directory {dir_path}: {e}")

    if not matched_directories:
        print("No matching directories found.")

def init_work_env(args, path,dist):

    if not os.path.exists(path):
        os.makedirs(path)

    out = os.path.join(args.path, "out")
    abi = os.path.join(args.path, "out_abi")

    if args.new.lower() != "false":
        print("this is a clean build,default delete out",args.new)
        if os.path.exists(out):
            print('\ndelete ', out)
            shutil.rmtree(out)

        if os.path.exists(abi):
            print('\ndelete ', abi)
            shutil.rmtree(abi)

        if os.path.exists(dist):
            print('\ndelete ', dist)
            shutil.rmtree(dist)

        patterns = [
            re.compile(r"android\d+-\d+\.\d+-\d+-\d+_r\d+\b"),
            re.compile(r"android\d+-\d+\.\d+-\d+-\d+_r\d+_abogki\d+\b")
        ]
        delete_matching_directories(args.path, patterns)

    if not os.path.exists(dist):
        os.makedirs(dist)

    if not os.path.exists(args.log):
        os.makedirs(args.log)

    os.chdir(path)
    #disable warning
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def run_repo_init_command(args, path, url, branch, manifest, ext):
    command = [repo_cmd, 'init']

    if url:
        command.extend(['-u', url])
    if branch:
        command.extend(['-b', branch])
    if manifest:
        command.extend(['-m', manifest])
    if ext:
        command.extend([ext])

    print('\ncommand',command)
    subprocess.run(command)

def run_repo_sync_command(args):
    command = "{} sync -c -j4 --prune --no-repo-verify --force-sync".format(repo_cmd)
    print('\ncommand',command)
    subprocess.run(command, shell=True)

def build_kernel_cmd(number):
    build_number = number.split("/")[0]
    command = []
    if build_number:
        os.environ['BUILD_NUMBER'] = "ogki" + build_number

    command.extend(['tools/bazel', 'run', '--config=release', '//common:kernel_aarch64_abi_dist'])
    print('command', command)
    print('number {} build_number {}'.format(number, build_number))
    env = os.environ.copy()  # 复制当前环境变量
    subprocess.run(command, env=env)  # 将复制的环境变量传递给subprocess.run()

def tar_dir_to_gz(zip, out_abi, log_file):

    if os.path.exists(zip):
        os.remove(zip)

    command = ['tar', '-czvf', zip, out_abi]
    print('\ncommand', command)
    with open(log_file, 'w') as f:
        result = subprocess.run(command, shell=False, stdout=f, stderr=subprocess.STDOUT)
        #print(result.stdout)

def copy_files_to_dist(src, file_paths, dst):
    if not os.path.exists(dst):
        os.makedirs(dst)
    for file_name, new_name in file_paths:
        src_path = os.path.join(src, file_name)
        dst_path = os.path.join(dst, new_name if new_name else file_name)
        if os.path.exists(src_path):
            print('copy {} to {}'.format(src_path, dst_path))
            shutil.copy(src_path, dst_path)
        else:
            print('{} not exist ignore'.format(src_path))

def get_build_info(output_path, repo_path='.'):
    try:
        result = subprocess.run(['repo', 'manifest', '-r', '-o', output_path], cwd=repo_path, stdout=subprocess.PIPE, check=True)
        return result.stdout.strip().decode('utf-8')
    except subprocess.CalledProcessError as e:
        return e.stderr.strip().decode('utf-8')

def default_download_gerrit_info(args, file_name, query):
    url = "{}/changes/?q={}".format(args.review_url, query)
    print('\nurl:', url)

    payload = {}
    files = {}
    headers = {}

    response = requests.get(url, headers=headers, data=payload, files=files, verify=False)
    print(response.text)
    with open(file_name, 'w') as file:
        file.write(response.text)

    print("\nResponse text already written to file:", file_name)

def session_download_gerrit_info(session, args, file_name, query):
    url = "{}/changes/?q={}".format(args.review_url, query)
    print('\nurl:', url)

    response = session.get(url, verify=False)
    #print(response.text)
    with open(file_name, 'w') as file:
        file.write(response.text)

    print("\nResponse text already written to file:", file_name)
    """"""

def find_default_user_info(args, email, remote_name, file_path='config_info.xlsx'):
    # 加载Excel工作簿
    print('\nfile_path', file_path)
    author_password = ""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 获取各列的索引
    column_names = [cell.value.strip() for cell in sheet[1]]
    #print(column_names)
    email_index = column_names.index("Committer_Email") + 1
    #print(email_index)
    password_index = column_names.index("password") + 1

    for row in range(2, sheet.max_row + 1):
        author_email = sheet.cell(row=row, column=email_index).value
        author_password = sheet.cell(row=row, column=password_index).value
        #print(author_email)
        #print(email)
        if author_email == email:
            break
    else:
        print("{} not found in the Excel file.".format(email))

    print(author_password)
    return author_password

def login_download_gerrit_info(args, file_name, query):
    # 你的Gerrit用户名和密码
    username = 'xxx'
    password = 'xxx'

    login_url = "{args.review_url}/login".format(args.review_url)
    print('\nlogin_url:', login_url)
    password = find_default_user_info(args, args.user, args.type, args.config_info)
    # 创建一个会话对象
    session = requests.Session()
    # 进行登录，获取cookie
    login_payload = {
        #'username': username,
        'username': args.user,
        'password': password
    }

    login_response = session.post(login_url, data=login_payload, verify=False)

    # 检查是否登录成功
    if login_response.status_code == 200:
        print("Login successful")
        session_download_gerrit_info(session, args, file_name, query)
    else:
        print("Login failed")

def download_gerrit_info(args, file_name, query):

    if args.type.lower() == "partner":
        print("to do later current use default")
        #login_download_gerrit_info(args, file_name, query)
    else:
        default_download_gerrit_info(args, file_name, query)

def filter_gerrit_info(input_file, output_file):
    with open(input_file, 'r') as f:
        lines = f.readlines()

    valid_jsons = []
    for line in lines:
        try:
            json.loads(line)
            valid_jsons.append(line)
        except json.JSONDecodeError:
            #print('\nignore invalid line \n\n{}\nin {}\n'.format(line, input_file))
            continue

    with open(output_file, 'w') as f:
        f.writelines(valid_jsons)

def parse_gerrit_info_from_json_to_excel(args, file_name, out_put, overwrite=True):
    extracted_data = []

    with open(file_name, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
            for item in data:

                branch = item['branch']
                virtual_id_number = str(item['virtual_id_number'])
                current_revision_number = str(item['current_revision_number'])

                last_two_digits = str(virtual_id_number)[-2:]
                refs = "refs/changes/{}/{}/{}".format(last_two_digits, virtual_id_number, current_revision_number)

                fetch_url = "{} {}".format(args.gerrit_url, refs)

                cherry_pick_command = "git fetch {} && git cherry-pick FETCH_HEAD ".format(fetch_url)
                checkout_command = "git fetch {} && git checkout FETCH_HEAD ".format(fetch_url)

                if args.type.lower() == "private":
                    push = "git push partner HEAD:refs/for/{}".format(branch)
                else:
                    push = "git push aosp HEAD:refs/for/{}".format(branch)

                extracted_item = {
                    'project': item['project'],
                    'branch': item['branch'],
                    'virtual_id_number': virtual_id_number,
                    'current_revision_number': current_revision_number,
                    'refs': refs,
                    'status': item['status'],
                    'cherry_pick_command': cherry_pick_command,
                    'checkout_command': checkout_command,
                    'push': push,
                    'change_id': item['change_id'],
                    'subject': item['subject']
                }
                extracted_data.append(extracted_item)
        except json.JSONDecodeError:
            print("by pass for error line", e)

    # 转换为DataFrame
    df = pd.DataFrame(extracted_data)

    if not overwrite:
        try:
            existing_df = pd.read_excel(out_put)
            merged_df = pd.concat([existing_df, df]).drop_duplicates()
            df = merged_df
        except FileNotFoundError:
            pass

    # 输出到Excel文件
    with pd.ExcelWriter(out_put, engine='openpyxl', mode='a' if not overwrite else 'w') as writer:
        df.to_excel(writer, index=False)
    print("data already write to {}".format(out_put))
    return df, df['status'].iloc[-1] if len(df) > 0 else None

def find_checkout_command_and_run(args, file_path, gerrit_id):
    """
    Reads an Excel file and finds the checkout command for a given gerrit_id.

    :param args: Arguments containing the common directory.
    :param file_path: Path to the Excel file.
    :param gerrit_id: gerrit_id to search for.
    :return: The last branch value corresponding to the matched gerrit_id.
    """
    branches = []
    try:
        # Read the Excel file
        df = pd.read_excel(file_path)

        # Ensure virtual_id_number is of type str
        df['virtual_id_number'] = df['virtual_id_number'].astype(str)

        # Ensure gerrit_id is also of type str
        gerrit_id = str(gerrit_id)

        # Filter the rows where virtual_id_number matches the given gerrit_id
        matched_rows = df[df['virtual_id_number'] == gerrit_id]

        # Check if any rows matched
        if not matched_rows.empty and len(matched_rows.columns) > 1:
            # Print the checkout_command column's content of matched rows
            for index, row in matched_rows.iterrows():
                command = row['checkout_command']
                branch = row['branch']
                # Print the checkout command and the corresponding branch
                print("Checkout Command: {}".format(command))
                print("Branch: {}".format(branch))
                # Execute the command
                current_path = os.getcwd()
                os.chdir(args.common)
                run_git_command(command)
                os.chdir(current_path)
                # Append the branch to the list
                branches.append(branch)
        else:
            print("No matching virtual_id_number found for gerrit_id: {}".format(gerrit_id))

    except Exception as e:
        print("An error occurred: {e}".format(e))

    # Print all matched branches
    for b in branches:
        print("Matched branch: {}".format(b))

    # Return the last branch if there are any branches matched
    return branches[-1] if branches else None

def tee_output_to_file_and_terminal(command, log_file):
    with open(log_file, 'w') as log:
        # 使用 Popen 而不是 run，因为 Popen 给我们提供了更多的控制
        process = subprocess.Popen(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            bufsize=1,  # 行缓冲
            universal_newlines=True  # 自动解码
        )

        # 逐行读取 stdout 和 stderr
        while True:
            output = process.stdout.readline()
            if output == '' and process.poll() is not None:
                break
            if output:
                sys.stdout.write(output)
                log.write(output)

        while True:
            error = process.stderr.readline()
            if error == '' and process.poll() is not None:
                break
            if error:
                sys.stderr.write(error)
                log.write(error)

        # 确保所有子进程流被关闭
        process.stdout.close()
        process.stderr.close()

        # 等待子进程的退出并得到返回码
        return_code = process.wait()
        return return_code

def upload_ogki_build(args, out, log_file):
    command = [
        "python3", get_tools_path("ogki_gki_artifactory.py"),
        "-t", "upload",
        "-k", "OGKI",
        "-i", out,
        "-o", args.artifactory
        #"-f"
    ]

    print('\ncommand:', command)

    try:
        return_code = tee_output_to_file_and_terminal(command, log_file)

        if return_code == 0:
            print('Successfully run command {}'.format(command))
        else:
            print('Failed to run command {}'.format(command))

    except Exception as e:
        print("run command error: {}".format(e))

def upload_ogki_buil_bak(args, out, log_file):
    command = [
        "python3", get_tools_path("ogki_gki_artifactory.py"),
        "-t", "upload",
        "-k", "OGKI",
        "-i", out,
        "-o", args.artifactory,
        "-f"
    ]
    print('\ncommand:', command)

    try:
        with open(log_file, 'w') as f:
            result = subprocess.run(command, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

            output = result.stdout.decode('utf-8')
            error = result.stderr.decode('utf-8')

            f.write(output)
            f.write(error)  # 写入错误日志

            if result.returncode == 0:
                print('Successfully run command {}'.format(command))
            else:
                print('Failed to run command {}'.format(command))
                print('Error:', error)

    except Exception as e:
        print("run command error: {}".format(e))

def git_format_patch(args, common, out, commit_hash):
    """
    Generate patches from the specified commit hash to the latest commit.

    :param common: The path to the git repository.
    :param out: The directory where the patch files will be generated.
    :param commit_hash: The commit hash to start generating patches from.
    """

    if  commit_hash:
        command = "git -C {} format-patch {}..HEAD -o {}".format(common, commit_hash, out)
    else:
        command = "git -C {} format-patch -{} -o {}".format(common, args.max, out)
    print('\nCommand:', command)
    try:
        subprocess.run(command, shell=True, check=True)
        print("Successfully generated patches.")
    except subprocess.CalledProcessError as e:
        print("Failed to execute command: {}".format(e))

def get_max_commits_list(repo_path, num_commits=3):
    # Git log command to get commits in a specified format
    cmd = "git -C {} log --pretty=format:'%H|%ae|%an|%ce|%cn|%s' --no-merges -{}".format(repo_path, num_commits)
    # Execute the command and get the output
    result = subprocess.check_output(cmd, shell=True).decode('utf-8')
    # Split the output into lines
    lines = result.strip().splitlines()
    # Split each line into fields
    data = [re.split(r'\|', line) for line in lines]

    # Predefine all expected columns
    columns = ['Commit_Hash', 'Author_Email', 'Author_Name', 'Committer_Email', 'Committer_Name', 'Subject',
               'project', 'branch', 'virtual_id_number', 'current_revision_number', 'refs',
               'status', 'cherry_pick_command', 'checkout_command', 'push', 'change_id', 'subject2']

    # Create an empty DataFrame with all possible columns
    df = pd.DataFrame(columns=columns)

    # Add rows to the DataFrame, filling missing columns with None or NaN
    rows = []
    for row in data:
        row_dict = {col: (row[index] if index < len(row) else None) for index, col in enumerate(columns[:len(row)])}
        rows.append(row_dict)

    # Use pd.concat to add rows
    df = pd.concat([df, pd.DataFrame(rows)], ignore_index=True, sort=False)

    return df

def write_commits_list_to_excel(df, output_file):
    df.to_excel(output_file, index=False)

def get_valid_commit_lists(args, file_path):
    merged_hash = ""
    # Load the Excel file
    df = pd.read_excel(file_path)

    # Find the column index for "project"
    project_col_index = df.columns.get_loc("project") + 1  # Excel column indices are 1-based

    # Find the index of the "Commit_Hash" column
    commit_hash_col_index = df.columns.get_loc("Commit_Hash")

    # Get the values of the "Commit_Hash" column
    commit_hash_column = df.iloc[:, commit_hash_col_index]

    # List to collect unprocessed commit hashes
    unprocessed_commit_hashes = []

    # Iterate over the items in the "Commit_Hash" column and process each commit_hash
    for index, commit_hash in enumerate(commit_hash_column):
        print(commit_hash)
        current_change_full_file = "current_change_{}_full.json".format(commit_hash)
        current_change_filtered_file = "current_change_{}_filtered.json".format(commit_hash)
        current_change_full_path = os.path.join(args.log, current_change_full_file)
        current_change_filtered_path = os.path.join(args.log, current_change_filtered_file)
        current_change_excel = os.path.join(args.log, "download_commit_file.xlsx")

        download_gerrit_info(args, current_change_full_path, commit_hash)
        filter_gerrit_info(current_change_full_path, current_change_filtered_path)
        df, status = parse_gerrit_info_from_json_to_excel(args, current_change_filtered_path, current_change_excel)

        # If status does not equal "MERGED", write the df values to the corresponding row starting from the project column
        if status is not None and status.upper() != "MERGED":
            print("Status is append, exiting...")
            # Determine the row index (1-based)
            row_index = index + 2  # Since we read from the first row, we need to add 2

            # Append unprocessed commit hashes
            unprocessed_commit_hashes.append(commit_hash)

            # Write the df values to the same row starting from the project column
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:  # Use append mode
                for col_index, value in enumerate(df.values[0], start=project_col_index):
                    writer.sheets['Sheet1'].cell(row=row_index, column=col_index, value=value)
        else:
            print("Status is merged, exiting...")
            merged_hash = commit_hash
            break

    # If the loop exits without setting merged_hash, it means no "MERGED" status was found.
    # Print and return the unprocessed commit hashes
    if not merged_hash:
        unprocessed_commit_hashes_str = ", ".join(unprocessed_commit_hashes)
        print("Unprocessed commit hashes:", unprocessed_commit_hashes_str)
        return unprocessed_commit_hashes_str
    else:
        print("Merged hash found:", merged_hash)
        return merged_hash

def set_git_config(name, email):
    # 设置全局用户名
    subprocess.run(["git", "config", "--global", "user.name", name])
    # 设置全局用户邮箱
    subprocess.run(["git", "config", "--global", "user.email", email])

def switch_to_branch(branch_name, remote_name='aosp'):
    # 获取当前分支名
    current_branch = subprocess.getoutput('git rev-parse --abbrev-ref HEAD')

    # 检查指定分支是否存在
    branches = subprocess.getoutput('git branch -a').split('\n')

    # 完全匹配的分支名
    matched_branches = [b.strip() for b in branches if b.strip().endswith('{}/{}'.format(remote_name, branch_name))]

    # 检查本地是否已经存在该分支
    local_branch_exists = branch_name in subprocess.getoutput('git branch')

    if matched_branches:
        if local_branch_exists:
            if current_branch != branch_name:
                subprocess.run(['git', 'checkout', branch_name])
        else:
            subprocess.run(['git', 'checkout', '-b', branch_name, 'remotes/{}/{}'.format(remote_name, branch_name)])
    else:
        print('Branch {} does not exist under remote {}.'.format(branch_name, remote_name))

def run_git_command(command):
    print('\ncommand', command)
    # 执行命令
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # 获取命令输出和错误信息
    output, error = process.communicate()
    # 返回命令的输出和错误信息
    if process.returncode == 0:
        print('Successfully run command  {}'.format(command))
        message = output.decode('utf-8')
        if message:
            print("Output:\n", message)
    else:
        print('Failed to  run command  {command}'.format(command))
        print("decode:\n", error.decode('utf-8'))

    return

def git_push(branch_name, remote_name='aosp'):
    command = ['git', 'push', remote_name, 'HEAD:refs/for/{}'.format(branch_name)]

    # 打印执行的命令
    print("Executing command: {}".format(command))

    result = subprocess.run(command)
    if result.returncode == 0:
        print('Successfully pushed to {} on branch {}'.format(remote_name, branch_name))
    else:
        print('Failed to push to {} on branch {}'.format(remote_name, branch_name))

def clear_gitcookies_script(shell_script):
    """
    Create a shell script to clear the contents of ~/.gitcookies.
    :param shell_script: The path to the shell script to generate.
    """
    with open(shell_script, 'w') as script:
        # Write shell script header
        script.write("#!/bin/bash\n\n")
        script.write("# This script clears the contents of ~/.gitcookies and executes a command\n\n")

        # Write command to clear .gitcookies
        script.write("echo 'Clearing the contents of ~/.gitcookies...'\n")
        script.write("> ~/.gitcookies\n")

    print("{} \ncreated to clear ~/.gitcookies and execute a command.".format(shell_script))

def append_command_to_script(shell_script, command):
    """
    Append a command to the shell script.
    :param shell_script: The path to the shell script.
    :param command: The command to append to the shell script.
    """
    with open(shell_script, 'a') as script:
        # Append the command
        script.write("\n# Execute the provided command\n")
        script.write("echo 'Executing command: {}'\n".format(command))
        script.write("{}\n".format(command))

def execute_shell_script(shell_script, output_file):
    """
    Execute the generated shell script.
    :param shell_script: The path to the shell script to execute.
    :param output_file: The file to write both stdout and stderr logs.
    """
    try:
        # Make the shell script executable
        subprocess.run("chmod +x {}".format(shell_script), shell=True, check=True)

        # Open the output_file in append mode
        with open(output_file, 'a') as file:
            # Execute the shell script and capture the output
            result = subprocess.run("{}".format(shell_script), shell=True, check=True, stdout=subprocess.PIPE)
            output = result.stdout.decode('utf-8')

            # Write the stdout and stderr to the file
            file.write("Shell script executed successfully:\n")
            file.write(output)

    except subprocess.CalledProcessError as e:
        # If an error occurs, write the stderr to the file
        with open(output_file, 'a') as file:
            file.write("Error occurred while executing shell script:\n")
            file.write(e.stderr)

def run_shell_command_with_script(args, shell_script, command):
    """
    Perform the task of clearing .gitcookies and executing a command using a shell script.
    :param shell_script: The path to the shell script to generate.
    :param command: The command to execute.
    """
    clear_gitcookies_script(shell_script)
    append_command_to_script(shell_script, command)
    execute_shell_script(shell_script, os.path.join(args.log, "custom_script.log"))

def default_set_gerrit_config_info(args, email, remote_name, file_path='config_info.xlsx'):
    # 加载Excel工作簿
    print('\nfile_path', file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 获取各列的索引
    column_names = [cell.value.strip() for cell in sheet[1]]
    #print(column_names)
    email_index = column_names.index("Committer_Email") + 1
    #print(email_index)
    aosp_index = column_names.index("aosp_Authenticate") + 1
    #print(aosp_index)
    partner_index = column_names.index("partner_Authenticate") + 1

    for row in range(2, sheet.max_row + 1):
        author_email = sheet.cell(row=row, column=email_index).value
        authenticate_aosp = sheet.cell(row=row, column=aosp_index).value
        authenticate_partner = sheet.cell(row=row, column=partner_index).value
        #print(author_email)
        #print(email)
        if author_email == email:
            if remote_name == 'partner':
                if authenticate_partner:
                    #print("Partner Authenticate for {}: {}".format(email, authenticate_partner))
                    run_shell_command_with_script(args, os.path.join(args.log, "custom_script.sh"), authenticate_partner)

                else:
                    print("No Partner Authenticate found for {}".format(email))
            else:
                if authenticate_aosp:
                    run_shell_command_with_script(args, os.path.join(args.log, "custom_script.sh"), authenticate_aosp)
                else:
                    print("No AOSP Authenticate found for {}".format(email))
            break
    else:
        print("{} not found in the Excel file.".format(email))

def rebase_valid_commits_list_and_push(args, file_path):
    # Read Excel file
    df = pd.read_excel(file_path)

    # Reverse order of rows
    df = df.iloc[::-1]

    # Iterate over rows
    for index, row in df.iterrows():
        if pd.notna(row['status']) and row['status'] != '':
            # Print each column value in the row
            #for col in df.columns:
            #    print("row {} {}: {}".format(index, col, row[col]))
            #print("\n")  # Newline for better readability
            Committer_Email = row['Committer_Email']
            Committer_Name = row['Committer_Name']
            branch = row['branch']
            set_git_config(Committer_Name, Committer_Email)
            default_set_gerrit_config_info(args, Committer_Email, args.type, args.config_info)
            switch_to_branch(branch, args.type)
            run_git_command(row['cherry_pick_command'])
            git_push(branch, args.type)
        #else:
        #    print("Empty status in the row {}".format(index))

def git_reset_and_pull(original_hash):
    """
    Function to reset git to a specific commit hash, pull the latest changes,
    and return the latest commit hash.
    :param original_hash: The hash to reset the git repository to.
    :return: The latest commit hash after pulling the latest changes.
    """
    try:
        # Step 2: Force reset to the specified commit hash
        result = subprocess.run(["git", "reset", "--hard", original_hash], check=True, stdout=subprocess.PIPE)
        print("Successfully reset to {}".format(original_hash))


        # Step 3: Pull the latest changes
        result = subprocess.run(["git", "pull"], check=True, stdout=subprocess.PIPE)
        print("Successfully pulled the latest changes")


        # Step 4: Get the latest commit hash
        result = subprocess.run(["git", "rev-parse", "HEAD"], check=True, stdout=subprocess.PIPE)
        latest_hash = result.stdout.strip().decode('utf-8')
        print("The latest commit hash is {}".format(latest_hash))

        return latest_hash

    except subprocess.CalledProcessError as e:
        print("An error occurred:\n{}".format(e.stderr))
        raise

def udpate_valid_commits_list_and_push(args, excel, branch, merged_hash):
    latest_hash = ""
    current_path = os.getcwd()
    os.chdir(args.common)

    print("switch branch: {}".format(branch))
    switch_to_branch(branch, args.type)
    # Example usage
    latest_hash = git_reset_and_pull(merged_hash)
    print("Updated to the latest hash: {}".format(latest_hash))

    rebase_valid_commits_list_and_push(args, excel)
    os.chdir(current_path)

    return latest_hash

def set_current_revision(args):

    current_path = os.getcwd()
    os.chdir(args.common)
    fetch_url = "{} {}".format(args.gerrit_url, args.refs)
    checkout_command = "git fetch {} && git checkout FETCH_HEAD ".format(fetch_url)
    #print('\ncommand', checkout_command)
    run_git_command(checkout_command)
    os.chdir(current_path)
    return ""

def set_latest_revision(args):

    full_download_json = os.path.join(args.log, "full_download.json")
    valid_download_json = os.path.join(args.log, "valid_download.json")
    valid_download_excel = os.path.join(args.log, "valid_download.xlsx")

    download_gerrit_info(args, full_download_json, args.gerrit_number)
    filter_gerrit_info(full_download_json, valid_download_json)
    parse_gerrit_info_from_json_to_excel(args, valid_download_json, valid_download_excel)
    branch = find_checkout_command_and_run(args, valid_download_excel, args.gerrit_number)
    print('\nbranch', branch)
    return branch

def pull_to_revision(args, max_commits_list_excel):
    merged_hash = ""
    branch = ""

    if args.auto:
        branch = set_latest_revision(args)
    else:
        branch = set_current_revision(args)

    if args.type.lower() == "aosp":
        df = get_max_commits_list(args.common, args.max)
        write_commits_list_to_excel(df, max_commits_list_excel)
        merged_hash = get_valid_commit_lists(args, max_commits_list_excel)

    print("merged_hash {}  branch: {}".format(merged_hash, branch))
    return merged_hash, branch

def run_repo_command(args):
    try:
        default_set_gerrit_config_info(args, args.user, args.type, args.config_info)
        run_repo_init_command(args, args.path, args.url, args.branch, args.manifest, args.ext)
        run_repo_sync_command(args)

    except Exception as e:
        # 打印异常信息（可选）
        print("repo error is: {}".format(e))
        sys.exit(1)
    return

def main():
    print('\nget input argc/argv')
    print('you can read this link for more help\n')

    args = parse_cmd_args()
    init_work_env(args, args.path, args.out)

    run_repo_command(args)

    max_commits_list_excel = os.path.join(args.log, "max_commits_list.xlsx")
    merged_hash, branch = pull_to_revision(args, max_commits_list_excel)

    if args.auto:
        merged_hash = udpate_valid_commits_list_and_push(args, max_commits_list_excel, branch, merged_hash)
    build_kernel_cmd(args.bug_number)
    git_format_patch(args, args.common, args.out, merged_hash)
    tar_log = os.path.join(args.log, "tar_log.txt")
    tar_dir_to_gz(args.zip, args.out_abi, tar_log)
    copy_files_to_dist(args.out_abi, file_paths, args.out)

    zip_log = os.path.join(args.out, "log.tar.gz")
    log_out = os.path.join(args.out, "log.txt")

    tar_dir_to_gz(zip_log, args.log, log_out)
    get_build_info(os.path.join(args.out, "BUILD_INFO.txt"))
    """"""
    upload_ogki_build(args, args.out, os.path.join(args.log, "upload_file.txt"))

if __name__ == "__main__":
    main()
